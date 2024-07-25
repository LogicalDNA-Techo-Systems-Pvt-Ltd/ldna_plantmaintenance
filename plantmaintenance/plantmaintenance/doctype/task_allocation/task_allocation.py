import frappe
import uuid
from frappe.model.document import Document
from frappe.utils import nowdate, getdate, add_days, add_years
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta
import calendar
import os
import openpyxl
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta


class TaskAllocation(Document):
    def on_update(self):
        if self.docstatus == 0:
            self.create_or_update_task_details()

    def check_tasks_generated(self):
        unique_keys = [detail.unique_key for detail in self.get('task_allocation_details')]
        return frappe.db.exists("Task Detail", {"unique_key": ["in", unique_keys]})

    def create_or_update_task_details(self):
        for detail in self.get('task_allocation_details'):
            existing_task = frappe.db.exists("Task Detail", {"unique_key": detail.unique_key})
            if not existing_task:
                self.create_task_detail(detail)
            else:
                self.update_task_detail(detail, existing_task)

    def create_task_detail(self, detail):
        parameter_info = self.get_parameter_info(detail.parameter)
        plan_start_date = detail.date
        require_time_hours = parameter_info.get('require_time')
        plan_end_date = (datetime.strptime(plan_start_date, '%Y-%m-%d') + timedelta(hours=require_time_hours)).strftime('%Y-%m-%d %H:%M:%S')

        task_detail = frappe.new_doc("Task Detail")
        task_detail.update({
            "unique_key": detail.unique_key,
            "task_allocation_id": self.name,
            "approver": frappe.session.user,
            "equipment_code": detail.equipment_code,
            "equipment_name": detail.equipment_name,
            "work_center": self.work_center,
            "plant_section": self.plant_section,
            "plan_start_date": plan_start_date,
            "plan_end_date": plan_end_date,
            "assigned_to": detail.assign_to,
            "activity": detail.activity,
            "parameter": detail.parameter,
            "parameter_type": parameter_info.get('parameter_type'),
            "minimum_value": parameter_info.get('minimum_value'),
            "maximum_value": parameter_info.get('maximum_value'),
            "require_time": parameter_info.get('require_time'),
            "values": ",".join(parameter_info.get('values', [])) if parameter_info.get('values') else None,
            "priority": detail.priority,
            "acceptance_criteria_for_list": parameter_info.get('acceptance_criteria_for_list'),
            "acceptance_criteria": parameter_info.get('acceptance_criteria'),

        })
        task_detail.insert(ignore_permissions=True)

    def update_task_detail(self, detail, task_name):
        task_detail_doc = frappe.get_doc("Task Detail", task_name)
        has_changes = False
        if task_detail_doc.assigned_to != detail.assign_to:
            task_detail_doc.assigned_to = detail.assign_to
            has_changes = True
        if task_detail_doc.priority != detail.priority:
            task_detail_doc.priority = detail.priority
            has_changes = True

        if has_changes:
            task_detail_doc.save(ignore_permissions=True)

    def get_parameter_info(self, parameter):
        parameter_doc = frappe.get_doc("Parameter", {"parameter": parameter})
        parameter_info = {
            "require_time": parameter_doc.require_time,
            "acceptance_criteria_for_list": parameter_doc.acceptance_criteria_for_list,            
            "acceptance_criteria": parameter_doc.acceptance_criteria,
            "parameter_type": parameter_doc.parameter_type,
            "minimum_value": parameter_doc.minimum_value if parameter_doc.parameter_type == "Numeric" else None,
            "maximum_value": parameter_doc.maximum_value if parameter_doc.parameter_type == "Numeric" else None,
            "values": parameter_doc.values.split(',') if parameter_doc.parameter_type == "List" else None
        }
        return parameter_info

@frappe.whitelist()
def check_tasks_generated(docname):
    if frappe.db.exists("Task Detail", {"task_allocation_id": docname}):
        return True
    return False

@frappe.whitelist()
def generate_tasks(docname):
    doc = frappe.get_doc("Task Allocation", docname)
    if doc:
        doc.create_or_update_task_details()
        return "Tasks have been generated successfully."
    else:
        return "Task Allocation document not found."


@frappe.whitelist()
def upload_tasks_excel_for_task_allocation(file,task_allocation_name):
    task_allocation_doc = frappe.get_doc("Task Allocation",task_allocation_name)    
   
    folder_path = ''
    actual_file_name = ''

    if file.startswith("/private/files/"):
        actual_file_name = file.replace("/private/files/", '')
        folder_path = os.path.join(os.path.abspath(frappe.get_site_path()), "private", "files")
    else:
        actual_file_name = file.replace("/files/", '')
        folder_path = os.path.join(os.path.abspath(frappe.get_site_path()), "public", "files")

    source_file = os.path.join(folder_path, actual_file_name)
 

    wb = openpyxl.load_workbook(source_file)
    sheet = wb.active

    headers = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)]
    
    task_allocation_doc.set("task_allocation_details", [])

    for row in range(2, sheet.max_row + 1):
       
        row_data = {}
        is_blank_row = True
        for col_num in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col_num).value
            if cell_value is not None and cell_value != '':
                is_blank_row = False
                break

        if is_blank_row:
            continue  
        
        for col_num in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col_num).value
            row_data[headers[col_num - 1]] = cell_value
            
        print(task_allocation_doc.as_dict())
        task_allocation_doc.append("task_allocation_details", {
            'equipment_code': row_data.get('Equipment Code'),
            'equipment_name': row_data.get('Equipment Name'),
            'activity': row_data.get('Activity'),
            'parameter': row_data.get('Parameter'),
            'frequency': row_data.get('Frequency'),
            'assign_to': row_data.get('Assign TO'),
            'date': row_data.get('Date'),
            'day': row_data.get('Day')
        })

        task_allocation_doc.save()
        

    return True


generated_unique_keys = set()

@frappe.whitelist()
def load_tasks(plant, location, functional_location, plant_section, work_center, end_date=None):
    global generated_unique_keys
    filters = {
        "plant": plant,
        "location": location,
        "functional_location": functional_location,
        "section": plant_section,
        "work_center": work_center
    }

    equipment_list = frappe.get_all('Equipment', filters=filters, fields=['equipment_code', 'equipment_name', 'activity_group'])
    setting_doc = frappe.get_single('Setting')
    start_date = getdate(setting_doc.start_date)
    today_date = getdate(nowdate())
    setting_end_date = getdate(setting_doc.end_date)
    end_date = getdate(end_date) if end_date else getdate(setting_doc.end_date)

    if end_date > setting_end_date:
        return frappe.msgprint("The end date must be less than the end date in the Setting doctype.")

    if not (start_date <= today_date <= end_date):
        return frappe.msgprint("Please ensure today's date is between the start date and end date.")

    tasks = []
    for equipment in equipment_list:
        activities = frappe.get_all('Activity CT', filters={'parent': equipment.activity_group}, fields=['activity'])

        for activity in activities:
            activity_details = frappe.get_doc('Activity', activity.activity)

            parameters = frappe.get_all('Parameter CT', filters={'parent': activity.activity}, fields=[
                'parameter', 'frequency', 'day_of_month', 'monday', 'tuesday', 'wednesday',
                'thursday', 'friday', 'saturday', 'sunday', 'date_of_year'
            ])

            for parameter in parameters:
                frequency = parameter.frequency
                dates = []

                if frequency == 'Daily':
                    dates = [add_days(today_date, i) for i in range(15) if today_date <= add_days(today_date, i) <= end_date]

                elif frequency == 'Weekly':
                    selected_days = []
                    if parameter.monday:
                        selected_days.append('Monday')
                    if parameter.tuesday:
                        selected_days.append('Tuesday')
                    if parameter.wednesday:
                        selected_days.append('Wednesday')
                    if parameter.thursday:
                        selected_days.append('Thursday')
                    if parameter.friday:
                        selected_days.append('Friday')
                    if parameter.saturday:
                        selected_days.append('Saturday')
                    if parameter.sunday:
                        selected_days.append('Sunday')

                    for day in selected_days:
                        current_date = today_date
                        while current_date.weekday() != list(calendar.day_name).index(day):
                            current_date += timedelta(days=1)
                        if today_date <= current_date <= end_date:
                            dates.append(current_date)

                        while current_date <= (today_date + timedelta(days=90)) and (today_date <= current_date <= end_date):
                            current_date += timedelta(weeks=1)
                            if today_date <= current_date <= end_date:
                                dates.append(current_date)

                elif frequency == 'Monthly':
                    day_of_month = parameter.day_of_month or 1
                    current_date = today_date.replace(day=day_of_month)
                    if current_date < today_date:
                        current_date += relativedelta(months=1)
                    dates = [current_date + relativedelta(months=i) for i in range(6) if today_date <= current_date + relativedelta(months=i) <= end_date]

                elif frequency == 'Yearly':
                    date_of_year = getdate(parameter.date_of_year)
                    if today_date <= date_of_year <= end_date:
                        years_range = range(today_date.year, end_date.year + 1)
                        dates = [date_of_year.replace(year=year) for year in years_range]

                for date in dates:
                    date_obj = getdate(date)
                    key_context = (equipment.equipment_code, activity_details.activity_name, parameter.parameter, date)
                    existing_key = next((key for key, context in generated_unique_keys if context == key_context), None)
                    if existing_key is None:
                        unique_key = 'lbvrq8' + str(uuid.uuid4())[:8]
                        generated_unique_keys.add((unique_key, key_context))
                    else:
                        unique_key = existing_key

                    task = {
                        'equipment_code': equipment.equipment_code,
                        'equipment_name': equipment.equipment_name,
                        'activity': activity_details.activity_name,
                        'parameter': parameter.parameter,
                        'frequency': frequency,
                        'date': date,
                        'day': calendar.day_name[date_obj.weekday()],
                        'unique_key': unique_key[:10]
                    }
                    tasks.append(task)


    return tasks


@frappe.whitelist()
def download_tasks_excel_for_task_allocation(tasks):
   tasks = frappe.parse_json(tasks)


   wb = Workbook()
   ws = wb.active
   ws.title = "Tasks"


   headers = ['Unique Key','Equipment Code', 'Equipment Name', 'Activity', 'Parameter', 'Frequency', 'Assign To', 'Date', 'Day','Priority']
   ws.append(headers)


   for task in tasks:
       task_date = getdate(task.get('date'))
       row = [
           task.get('unique_key'),
           task.get('equipment_code'),
           task.get('equipment_name'),
           task.get('activity'),
           task.get('parameter'),
           task.get('frequency'),
           task.get('assign_to'),
           task_date.strftime('%Y-%m-%d') if task_date else None,
           task.get('day'),
            task.get('priority')
       ]
       ws.append(row)


   virtual_workbook = BytesIO()
   wb.save(virtual_workbook)
   virtual_workbook.seek(0)
   file_data = virtual_workbook.read()


   file_name = f'Task_Allocation_{nowdate()}.xlsx'
   file_doc = frappe.get_doc({
       'doctype': 'File',
       'file_name': file_name,
       'content': file_data,
       'is_private': 0
   })
   file_name = f'Task_Detail_{nowdate()}.xlsx'
   file_doc = frappe.get_doc({
       'doctype': 'File',
       'file_name': file_name,
       'content': file_data,
       'is_private': 0
   })
   file_doc.save(ignore_permissions=True)


   return file_doc.file_url


@frappe.whitelist()
def upload_tasks_excel_for_task_allocation(file, task_allocation_name):
   task_allocation_doc = frappe.get_doc("Task Allocation", task_allocation_name)   


   folder_path = ''
   actual_file_name = ''


   if file.startswith("/private/files/"):
       actual_file_name = file.replace("/private/files/", '')
       folder_path = os.path.join(os.path.abspath(frappe.get_site_path()), "private", "files")
   else:
       actual_file_name = file.replace("/files/", '')
       folder_path = os.path.join(os.path.abspath(frappe.get_site_path()), "public", "files")


   source_file = os.path.join(folder_path, actual_file_name)


   wb = load_workbook(source_file)
   sheet = wb.active


   headers = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)]


   task_allocation_doc.set("task_allocation_details", [])


   for row in range(2, sheet.max_row + 1):
       row_data = {}
       is_blank_row = True
       for col_num in range(1, sheet.max_column + 1):
           cell_value = sheet.cell(row=row, column=col_num).value
           if cell_value is not None and cell_value != '':
               is_blank_row = False
               break


       if is_blank_row:
           continue 


       for col_num in range(1, sheet.max_column + 1):
           cell_value = sheet.cell(row=row, column=col_num).value
           row_data[headers[col_num - 1]] = cell_value


       task_allocation_doc.append("task_allocation_details", {
           'equipment_code': row_data.get('Equipment Code'),
           'equipment_name': row_data.get('Equipment Name'),
           'activity': row_data.get('Activity'),
           'parameter': row_data.get('Parameter'),
           'frequency': row_data.get('Frequency'),
           'date': row_data.get('Date'),
           'assign_to': row_data.get('Assign To'),
           'priority': row_data.get('Priority'),
           'day': row_data.get('Day'),
           'unique_key': row_data.get('Unique Key')
       })


   task_allocation_doc.save(ignore_permissions=True)
  
   return {"message": "Excel import successful!"}
