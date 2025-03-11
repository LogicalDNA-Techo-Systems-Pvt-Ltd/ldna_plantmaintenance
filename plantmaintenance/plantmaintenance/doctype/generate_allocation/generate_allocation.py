# Copyright (c) 2025, LogicalDNA and contributors
# For license information, please see license.txt

from frappe.model.document import Document

class GenerateAllocation(Document):
	pass

import frappe
import uuid
from frappe.model.document import Document
from frappe.utils import nowdate, getdate, add_days
import calendar
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta
import os
import openpyxl
import base64
import io
from frappe.exceptions import TimestampMismatchError
from datetime import timedelta
from frappe.utils import getdate, add_days
from dateutil.relativedelta import relativedelta

@frappe.whitelist()
def load_tasks(plant, location, plant_section, work_center, end_date=None, equipment_list=None):
    today = getdate()

    if not end_date:
        frappe.throw("Please provide an End Date.")

    end_date = getdate(end_date)

    if equipment_list:
        if isinstance(equipment_list, str):
            try:
                equipment_list = frappe.parse_json(equipment_list)
            except Exception:
                frappe.throw("Invalid Equipment List format.")
    else:
        equipment_list = frappe.db.get_list(
            "Task Detail",
            filters={},
            pluck="equipment_code",
            distinct=True
        )
        if not equipment_list:
            frappe.throw("No tasks exist for any equipment.")

    frappe.logger().info(f"Final Equipment List for Task Generation: {equipment_list}")

    new_tasks = []

    for equipment_code in equipment_list:
        existing_tasks = frappe.get_all(
            "Task Detail",
            filters={"equipment_code": equipment_code},
            fields=["equipment_code", "activity", "activity_group", "parameter", "frequency", "plan_start_date", "type"]
        )

        for task in existing_tasks:
            if task.get("type") and task["type"] != "Preventive":
                continue
            
            if not task.get("parameter"):
                frappe.logger().warning(f"Skipping task for {task['equipment_code']} - Parameter is missing.")
                continue

            frequency = task["frequency"]
            plan_start_date = getdate(task["plan_start_date"])
            next_due_date = plan_start_date

            while next_due_date <= end_date:
                date_obj = getdate(next_due_date)
                unique_key = f"task_{task['equipment_code']}_{task['activity']}_{task['parameter']}_{next_due_date}"

                if frappe.db.exists("Task Detail", {
                    "equipment_code": task['equipment_code'],
                    "activity": task['activity'],
                    "parameter": task['parameter'],
                    "frequency": task['frequency'],
                    "plan_start_date": next_due_date
                }):
                    next_due_date = calculate_next_due_date(next_due_date, frequency)
                    continue 

                equipment_fields = ["equipment_name", "sub_section", "old_tag_dcs", "description"]
                equipment_name, sub_section, old_tag_dcs, description = frappe.db.get_value("Equipment", task['equipment_code'], equipment_fields)

                parameter_doc = frappe.get_doc("Parameter", task['parameter']) if task["parameter"] else None
                parameter_type = parameter_doc.parameter_type if parameter_doc else None
                minimum_value = parameter_doc.minimum_value if parameter_doc else None
                maximum_value = parameter_doc.maximum_value if parameter_doc else None
                standard_value = parameter_doc.standard_value if parameter_doc else None

                new_task = {
                    'equipment_code': task['equipment_code'],
                    'equipment_name': equipment_name,  
                    'activity_group': task['activity_group'], 
                    'activity': task['activity'],
                    'parameter': task['parameter'],
                    'frequency': frequency,
                    'date': next_due_date,
                    'day': calendar.day_name[date_obj.weekday()],
                    'unique_key': unique_key[:10],
                    'sub_section': sub_section,
                    'old_tag_dcs': old_tag_dcs,
                    'description': description,
                    'parameter_type': parameter_type,
                    'minimum_value': minimum_value,
                    'maximum_value': maximum_value,
                    'standard_value': standard_value
                }

                new_tasks.append(new_task)

                task_detail = frappe.new_doc("Task Detail")
                task_detail.update({
                    "approver": frappe.session.user,
                    "equipment_code": new_task['equipment_code'],
                    "equipment_name": new_task['equipment_name'],  
                    "activity_group": new_task['activity_group'],  
                    "activity": new_task['activity'],
                    "parameter": new_task['parameter'],
                    "frequency": new_task['frequency'],
                    "plan_start_date": new_task['date'],
                    "day": new_task['day'],
                    "date": new_task['date'],
                    "unique_key": new_task['unique_key'],
                    "location": location,
                    "section": plant_section,
                    "old_tag_dcs": new_task["old_tag_dcs"],
                    "sub_section": new_task["sub_section"],
                    "description": new_task["description"],
                    "parameter_type": new_task["parameter_type"],
                    "minimum_value": new_task["minimum_value"],
                    "maximum_value": new_task["maximum_value"],
                    "standard_value": new_task["standard_value"]
                })
                task_detail.insert(ignore_permissions=True)

                next_due_date = calculate_next_due_date(next_due_date, frequency)

    frappe.logger().info(f"Total New Tasks Generated: {len(new_tasks)}")
    
    return new_tasks


def calculate_next_due_date(current_date, frequency):
    """Helper function to calculate the next due date based on frequency."""
    if frequency == 'Daily':
        return add_days(current_date, 1)
    elif frequency == 'Weekly':
        return add_days(current_date, 7)
    elif frequency == 'By Weekly':
        return add_days(current_date, 14)
    elif frequency == 'Monthly':
        return current_date + relativedelta(months=1)
    elif frequency == 'Quarterly':
        return current_date + relativedelta(months=3)
    elif frequency == 'Half-Yearly':
        return current_date + relativedelta(months=6)
    elif frequency == 'Yearly':
        return current_date + relativedelta(years=1)
    elif frequency == 'Two-Yearly':
        return current_date + relativedelta(years=2)
    elif frequency == 'Five-Yearly': 
        return current_date + relativedelta(years=5)
    
    return current_date



@frappe.whitelist()
def download_tasks_excel_for_allocation(tasks):
    doc = frappe.get_doc("Generate Allocation")
    tasks = frappe.parse_json(tasks)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"


    headers = ['Unique Key','Equipment Code', 'Equipment Name', 'Activity Group','Activity', 'Parameter', 'Frequency', 'Assign To', 'Date', 'Day','Priority']
    ws.append(headers)


    for task in tasks:
        task_date = getdate(task.get('date'))
        row = [
            task.get('unique_key'),
            task.get('equipment_code'),
            task.get('equipment_name'),
            task.get('activity_group'),
            task.get('activity'),
            task.get('parameter'),
            task.get('frequency'),
            task.get('assign_to'),
            task_date.strftime('%Y-%m-%d') if task_date else None,
            task.get('day'),
            task.get('priority'),
        ]
        ws.append(row)


    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    file_data = virtual_workbook.read()


    file_name = f'Task_Allocation_{nowdate()}.xlsx'
    file_doc = frappe.get_doc({
        'doctype': 'File',
        'file_name': file_name,
        'content': file_data,
        'is_private': 0
    })
    file_name = f'Task_Detail_{nowdate()}.xlsx'
    file_doc = frappe.get_doc({
        'doctype': 'File',
        'file_name': file_name,
        'content': file_data,
        'is_private': 0
    })
    file_doc.save(ignore_permissions=True)


    return file_doc.file_url
 

@frappe.whitelist()
def upload_tasks_excel_for_allocation(file, allocation_name):
    folder_path = ''
    actual_file_name = ''

    if file.startswith("/private/files/"):
        actual_file_name = file.replace("/private/files/", '')
        folder_path = os.path.join(os.path.abspath(frappe.get_site_path()), "private", "files")
    else:
        actual_file_name = file.replace("/files/", '')
        folder_path = os.path.join(os.path.abspath(frappe.get_site_path()), "public", "files")

    source_file = os.path.join(folder_path, actual_file_name)
    wb = load_workbook(source_file)
    sheet = wb.active

    headers = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)]
    allocation_details = []

    missing_assign_to_count = 0
    mismatch_people_count = 0

    for row in range(2, sheet.max_row + 1):
        row_data = {}
        is_blank_row = True
        for col_num in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col_num).value
            if cell_value is not None and cell_value != '':
                is_blank_row = False
                break

        if is_blank_row:
            continue

        for col_num in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col_num).value
            row_data[headers[col_num - 1]] = cell_value

        assign_to = row_data.get('Assign To')
        parameter_doc = frappe.get_doc("Parameter", row_data.get('Parameter'))

        if assign_to is None or assign_to.strip() == '':
            missing_assign_to_count += 1
        elif parameter_doc.number_of_maintenance_person != len(assign_to.split(',')):
            mismatch_people_count += 1

        allocation_details.append({
            'equipment_code': row_data.get('Equipment Code'),
            'equipment_name': row_data.get('Equipment Name'),
            'activity_group': row_data.get('Activity Group'),
            'activity': row_data.get('Activity'),
            'parameter': row_data.get('Parameter'),
            'frequency': row_data.get('Frequency'),
            'date': getdate(row_data.get('Date')),
            'assign_to': assign_to,
            'priority': row_data.get('Priority'),
            'day': row_data.get('Day')
        })

    
    for detail in allocation_details:
        filters = {
            'equipment_code': detail['equipment_code'],
            'activity': detail['activity'],
            'parameter': detail['parameter'],
            'plan_start_date': detail['date']
        }

        task_details = frappe.get_all('Task Detail', filters=filters, fields=['name'])

        if not task_details:
            frappe.log_error(f"No tasks found for filters: {filters}", "Task Update Error")
            continue

        for task in task_details:
            try:
                task_detail = frappe.get_doc('Task Detail', task.name)
                task_detail.assigned_to = detail['assign_to']
                task_detail.priority = detail['priority']
                task_detail.save(ignore_permissions=True)
            except Exception as e:
                frappe.log_error(f"Error updating task {task.name}: {str(e)}", "Task Update Error")

    error_message = ""
    if missing_assign_to_count > 0:
        error_message += f"{missing_assign_to_count} rows are missing 'Assign To' person. "
    if mismatch_people_count > 0:
        error_message += f"{mismatch_people_count} rows have a mismatch in the number of people assigned."

    if error_message:
        frappe.msgprint(error_message)

    return {"message": "Excel import successful with warnings!" if error_message else "Excel import successful!", "allocation_details": allocation_details}

