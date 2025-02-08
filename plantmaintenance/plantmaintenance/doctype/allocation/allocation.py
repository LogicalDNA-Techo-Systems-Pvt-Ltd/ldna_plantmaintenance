import frappe
import uuid
from frappe.model.document import Document
from frappe.utils import nowdate, getdate, add_days
import calendar
from openpyxl import Workbook, load_workbook
from io import BytesIO
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta
import os
import openpyxl
import base64
import io
from frappe.exceptions import TimestampMismatchError

class Allocation(Document):
    pass

generated_unique_keys = set()

# @frappe.whitelist()
# def load_tasks(plant, location, plant_section, work_center, start_date=None, end_date=None, equipment=None):
#     global generated_unique_keys

#     today = getdate()
#     if start_date and end_date and getdate(start_date) > getdate(end_date):
#         frappe.throw("Start Date must be less than or equal to End Date.")

#     if start_date and getdate(start_date) < today:
#         frappe.throw("Start Date cannot be before today's date.")

#     settings_doc = frappe.get_single('Settings')
#     start_date = getdate(start_date) if start_date else max(today, getdate(settings_doc.start_date))
#     end_date = getdate(end_date) if end_date else getdate(settings_doc.end_date)

#     current_user = frappe.session.user
#     user_work_center = frappe.get_value('User Work Center', {'user': current_user}, 'name')
#     assigned_work_centers = frappe.get_all('Work Center CT', filters={'parent': user_work_center}, pluck='work_center')
    
#     if work_center not in assigned_work_centers or not assigned_work_centers or not user_work_center:
#         return frappe.msgprint(f"You are not assigned to {work_center} work center")

#     filters = {
#         "plant": plant,
#         "location": location,
#         "section": plant_section,
#         "work_center": work_center,
#         "on_scrap": 0, 
#         "activity_group_active": 1 
#     }
#     if equipment:
#         filters["equipment_code"] = equipment

#     equipment_list = frappe.get_all('Equipment', filters=filters, fields=['equipment_code', 'equipment_name', 'activity_group'])
#     if not equipment_list:
#         return frappe.msgprint("No equipment found for the provided filters.")

#     tasks = []
    
#     for equipment_item in equipment_list:
#         if not equipment_item.activity_group:
#             continue

#         activities = frappe.get_all('Activity CT', filters={'parent': equipment_item.activity_group}, fields=['activity'])
#         if not activities:
#             continue

#         for activity in activities:
#             activity_details = frappe.get_doc('Activity', activity.activity)
#             parameters = frappe.get_all('Parameter CT', filters={'parent': activity.activity}, fields=['parameter', 'frequency'])
#             if not parameters:
#                 continue

#             for parameter in parameters:
#                 frequency = parameter.frequency
#                 dates = []
                
#                 if frequency == 'Daily':
#                     dates = [add_days(start_date, i) for i in range((end_date - start_date).days + 1)]
                
#                 elif frequency == 'Weekly':
#                     current_date = start_date
#                     while current_date <= end_date:
#                         dates.append(current_date)
#                         current_date += timedelta(weeks=1)
                
#                 elif frequency == 'By Weekly':
#                     current_date = start_date
#                     while current_date <= end_date:
#                         dates.append(current_date)
#                         current_date += timedelta(weeks=2)
                
#                 elif frequency == 'Monthly':
#                     current_date = start_date
#                     while current_date <= end_date:
#                         dates.append(current_date)
#                         current_date += relativedelta(months=1)
                
#                 elif frequency == 'Quarterly':
#                     current_date = start_date
#                     while current_date <= end_date:
#                         dates.append(current_date)
#                         current_date += relativedelta(months=3)
                
#                 elif frequency == 'Half-Yearly':
#                     current_date = start_date
#                     while current_date <= end_date:
#                         dates.append(current_date)
#                         current_date += relativedelta(months=6)
                
#                 elif frequency == 'Yearly':
#                     current_date = start_date
#                     while current_date <= end_date:
#                         dates.append(current_date)
#                         current_date += relativedelta(years=1)
                
#                 for date in dates:
#                     date_obj = getdate(date)
#                     unique_key = 'lbvrq8' + str(uuid.uuid4())[:8]
                    
#                     task = {
#                         'equipment_code': equipment_item.equipment_code,
#                         'equipment_name': equipment_item.equipment_name,
#                         'activity_group': equipment_item.activity_group,
#                         'activity': activity_details.activity_name,
#                         'parameter': parameter.parameter,
#                         'frequency': frequency,
#                         'date': date,
#                         'day': calendar.day_name[date_obj.weekday()],
#                         'unique_key': unique_key[:10]
#                     }
                    
#                     tasks.append(task)
                    
#                     if not frappe.db.exists('Task Detail', {
#                         'equipment_code': task['equipment_code'],
#                         'activity': task['activity'],
#                         'parameter': task['parameter'],
#                         'frequency': task['frequency'],
#                         'plan_start_date': task['date']
#                     }):
#                         task_detail = frappe.new_doc("Task Detail")
#                         task_detail.update({
#                             "approver": frappe.session.user,
#                             "equipment_code": task['equipment_code'],
#                             "equipment_name": task['equipment_name'],
#                             "activity_group": task['activity_group'],
#                             "work_center": work_center,
#                             "plant_section": plant_section,
#                             "location": location,
#                             "plan_start_date": task['date'],
#                             "activity": task['activity'],
#                             "parameter": task['parameter'],
#                             "frequency": task['frequency'],
#                             "day": task['day'],
#                             "date": task['date'],
#                             "unique_key": task['unique_key']
#                         })
#                         task_detail.insert(ignore_permissions=True)
    
#     if not tasks:
#         return frappe.msgprint("No tasks found for the provided filters.")
    
#     return tasks

import frappe
import uuid
import calendar
from datetime import timedelta
from frappe.utils import getdate, add_days
from dateutil.relativedelta import relativedelta

@frappe.whitelist()
def load_tasks(plant, location, plant_section, work_center, start_date=None, end_date=None, equipment=None):
    today = getdate()

    if start_date and end_date and getdate(start_date) > getdate(end_date):
        frappe.throw("Start Date must be less than or equal to End Date.")

    if start_date and getdate(start_date) < today:
        frappe.throw("Start Date cannot be before today's date.")

    settings_doc = frappe.get_single('Settings')
    start_date = getdate(start_date) if start_date else max(today, getdate(settings_doc.start_date))
    end_date = getdate(end_date) if end_date else getdate(settings_doc.end_date)

    current_user = frappe.session.user
    
    # user_work_center_doc = frappe.get_all(
    #     "User Work Center",
    #     filters={"user": current_user},
    #     fields=["name", "equipment_group"]
    # )
    user_work_center_doc = frappe.get_all(
        "User Work Center",
        filters={"user": current_user},
        fields=["name"]
    )

    if not user_work_center_doc:
        return frappe.msgprint("You are not assigned to any Work Center or Equipment Group.")

    user_work_center_name = user_work_center_doc[0]["name"]
    # assigned_equipment_group = user_work_center_doc[0].get("equipment_group")
    
    # work_centers = frappe.get_all(
    #     "Work Center CT",
    #     filters={"parent": user_work_center_doc[0]["name"]},
    #     pluck="work_center"
    # )

    # if not assigned_equipment_group and not work_centers:
    #     return frappe.msgprint("You are not assigned to any Work Center and Equipment Group.")

    # if not assigned_equipment_group:
    #     return frappe.msgprint("You are not assigned to any Equipment Group.")

    # if not work_centers:
    #     return frappe.msgprint("You are not assigned to any Work Center.")

    # if work_center not in work_centers:
    #     return frappe.msgprint(f"You are assigned to Equipment Group {assigned_equipment_group}, but not to the {work_center} Work Center.")

    assigned_work_centers = frappe.get_all(
        "Work Center CT",
        filters={"parent": user_work_center_name},
        pluck="work_center"
    )

    assigned_equipment_groups = frappe.get_all(
        "Equipment Group CT",
        filters={"parent": user_work_center_name},
        pluck="equipment_group"
    )

    if not assigned_work_centers:
        return frappe.msgprint("You are not assigned to any Work Center.")

    if not assigned_equipment_groups:
        return frappe.msgprint("You are not assigned to any Equipment Group.")

    if work_center not in assigned_work_centers:
        return frappe.msgprint(f"You are assigned to Work Centers {', '.join(assigned_work_centers)}, but not to {work_center}.")

    filters = {
        "plant": plant,
        "location": location,
        "section": plant_section,
        "work_center": work_center,
        "on_scrap": 0,
        "activity_group_active": 1,
        "equipment_group": ["in", assigned_equipment_groups]
    }

    if equipment:
        filters["equipment_code"] = equipment

    equipment_list = frappe.get_all('Equipment', filters=filters, fields=['equipment_code', 'equipment_name', 'activity_group', 'equipment_group'])
    
    if not equipment_list:
        return frappe.msgprint("No equipment found for the provided filters.")

    tasks = []

    for equipment_item in equipment_list:
        # if equipment_item.equipment_group != assigned_equipment_group:
        #     frappe.throw(f"Equipment Group mismatch: You are assigned to Equipment Group '{assigned_equipment_group}', but equipment '{equipment_item.equipment_code}' belongs to '{equipment_item.equipment_group}'.")

        if not equipment_item.activity_group:
            continue

        activities = frappe.get_all('Activity CT', filters={'parent': equipment_item.activity_group}, fields=['activity'])

        if not activities:
            continue

        for activity in activities:
            activity_details = frappe.get_doc('Activity', activity.activity)
            parameters = frappe.get_all('Parameter CT', filters={'parent': activity.activity}, fields=['parameter', 'frequency'])

            if not parameters:
                continue

            for parameter in parameters:
                frequency = parameter.frequency
                dates = []

                if frequency == 'Daily':
                    dates = [add_days(start_date, i) for i in range((end_date - start_date).days + 1)]
                elif frequency == 'Weekly':
                    current_date = start_date
                    while current_date <= end_date:
                        dates.append(current_date)
                        current_date += timedelta(weeks=1)
                elif frequency == 'By Weekly':
                    current_date = start_date
                    while current_date <= end_date:
                        dates.append(current_date)
                        current_date += timedelta(weeks=2)
                elif frequency == 'Monthly':
                    current_date = start_date
                    while current_date <= end_date:
                        dates.append(current_date)
                        current_date += relativedelta(months=1)
                elif frequency == 'Quarterly':
                    current_date = start_date
                    while current_date <= end_date:
                        dates.append(current_date)
                        current_date += relativedelta(months=3)
                elif frequency == 'Half-Yearly':
                    current_date = start_date
                    while current_date <= end_date:
                        dates.append(current_date)
                        current_date += relativedelta(months=6)
                elif frequency == 'Yearly':
                    current_date = start_date
                    while current_date <= end_date:
                        dates.append(current_date)
                        current_date += relativedelta(years=1)

                for date in dates:
                    date_obj = getdate(date)
                    unique_key = 'lbvrq8' + str(uuid.uuid4())[:8]

                    task = {
                        'equipment_code': equipment_item.equipment_code,
                        'equipment_name': equipment_item.equipment_name,
                        'activity_group': equipment_item.activity_group,
                        'activity': activity_details.activity_name,
                        'parameter': parameter.parameter,
                        'frequency': frequency,
                        'date': date,
                        'day': calendar.day_name[date_obj.weekday()],
                        'unique_key': unique_key[:10]
                    }

                    tasks.append(task)

                    if not frappe.db.exists('Task Detail', {
                        'equipment_code': task['equipment_code'],
                        'activity': task['activity'],
                        'parameter': task['parameter'],
                        'frequency': task['frequency'],
                        'plan_start_date': task['date']
                    }):
                        task_detail = frappe.new_doc("Task Detail")
                        task_detail.update({
                            "approver": frappe.session.user,
                            "equipment_code": task['equipment_code'],
                            "equipment_name": task['equipment_name'],
                            "activity_group": task['activity_group'],
                            "work_center": work_center,
                            "plant_section": plant_section,
                            "location": location,
                            "plan_start_date": task['date'],
                            "activity": task['activity'],
                            "parameter": task['parameter'],
                            "frequency": task['frequency'],
                            "day": task['day'],
                            "date": task['date'],
                            "unique_key": task['unique_key']
                        })
                        task_detail.insert(ignore_permissions=True)

    if not tasks:
        return frappe.msgprint("No tasks found for the provided filters.")
    
    return tasks


@frappe.whitelist()
def download_tasks_excel_for_allocation(tasks):
    doc = frappe.get_doc("Allocation")
    tasks = frappe.parse_json(tasks)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"


    headers = ['Unique Key','Equipment Code', 'Equipment Name', 'Activity Group','Activity', 'Parameter', 'Frequency', 'Assign To', 'Date', 'Day','Priority']
    ws.append(headers)


    for task in tasks:
        task_date = getdate(task.get('date'))
        row = [
            task.get('unique_key'),
            task.get('equipment_code'),
            task.get('equipment_name'),
            task.get('activity_group'),
            task.get('activity'),
            task.get('parameter'),
            task.get('frequency'),
            task.get('assign_to'),
            task_date.strftime('%Y-%m-%d') if task_date else None,
            task.get('day'),
            task.get('priority'),
        ]
        ws.append(row)


    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    file_data = virtual_workbook.read()


    file_name = f'Task_Allocation_{nowdate()}.xlsx'
    file_doc = frappe.get_doc({
        'doctype': 'File',
        'file_name': file_name,
        'content': file_data,
        'is_private': 0
    })
    file_name = f'Task_Detail_{nowdate()}.xlsx'
    file_doc = frappe.get_doc({
        'doctype': 'File',
        'file_name': file_name,
        'content': file_data,
        'is_private': 0
    })
    file_doc.save(ignore_permissions=True)


    return file_doc.file_url
 

@frappe.whitelist()
def upload_tasks_excel_for_allocation(file, allocation_name):
    folder_path = ''
    actual_file_name = ''

    if file.startswith("/private/files/"):
        actual_file_name = file.replace("/private/files/", '')
        folder_path = os.path.join(os.path.abspath(frappe.get_site_path()), "private", "files")
    else:
        actual_file_name = file.replace("/files/", '')
        folder_path = os.path.join(os.path.abspath(frappe.get_site_path()), "public", "files")

    source_file = os.path.join(folder_path, actual_file_name)
    wb = load_workbook(source_file)
    sheet = wb.active

    headers = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)]
    allocation_details = []

    missing_assign_to_count = 0
    mismatch_people_count = 0

    for row in range(2, sheet.max_row + 1):
        row_data = {}
        is_blank_row = True
        for col_num in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col_num).value
            if cell_value is not None and cell_value != '':
                is_blank_row = False
                break

        if is_blank_row:
            continue

        for col_num in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col_num).value
            row_data[headers[col_num - 1]] = cell_value

        assign_to = row_data.get('Assign To')
        parameter_doc = frappe.get_doc("Parameter", row_data.get('Parameter'))

        if assign_to is None or assign_to.strip() == '':
            missing_assign_to_count += 1
        elif parameter_doc.number_of_maintenance_person != len(assign_to.split(',')):
            mismatch_people_count += 1

        allocation_details.append({
            'equipment_code': row_data.get('Equipment Code'),
            'equipment_name': row_data.get('Equipment Name'),
            'activity_group': row_data.get('Activity Group'),
            'activity': row_data.get('Activity'),
            'parameter': row_data.get('Parameter'),
            'frequency': row_data.get('Frequency'),
            'date': getdate(row_data.get('Date')),
            'assign_to': assign_to,
            'priority': row_data.get('Priority'),
            'day': row_data.get('Day')
        })

    
    for detail in allocation_details:
        filters = {
            'equipment_code': detail['equipment_code'],
            'activity': detail['activity'],
            'parameter': detail['parameter'],
            'plan_start_date': detail['date']
        }

        task_details = frappe.get_all('Task Detail', filters=filters, fields=['name'])

        if not task_details:
            frappe.log_error(f"No tasks found for filters: {filters}", "Task Update Error")
            continue

        for task in task_details:
            try:
                task_detail = frappe.get_doc('Task Detail', task.name)
                task_detail.assigned_to = detail['assign_to']
                task_detail.priority = detail['priority']
                task_detail.save(ignore_permissions=True)
            except Exception as e:
                frappe.log_error(f"Error updating task {task.name}: {str(e)}", "Task Update Error")

    error_message = ""
    if missing_assign_to_count > 0:
        error_message += f"{missing_assign_to_count} rows are missing 'Assign To' person. "
    if mismatch_people_count > 0:
        error_message += f"{mismatch_people_count} rows have a mismatch in the number of people assigned."

    if error_message:
        frappe.msgprint(error_message)

    return {"message": "Excel import successful with warnings!" if error_message else "Excel import successful!", "allocation_details": allocation_details}

